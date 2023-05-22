#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Aug 18 10:51:33 2021

Description:
    1. Read files in folder where each file contains one-to-all gene interactions and scores.
        i) Only keep longest-sequenced gene isoform if different isoforms exist in files
    2. For each file:
        i) Sort interactors in descending order by score (last column in each file)
        ii) Pull top X (20 for soy-pathogen, 40 for soy-soy) scoring interactors for given gene file (# top scorers is an option)
        iii) Calculate % of top scorers that include genes of interest
    3. Add top scorers to Excel under gene column
        i) Format and save as Excel, highlighting any genes of interest found in each column

Usage:
    Open terminal (command-line interface)
    Change to directory/folder where extract_top_genes.py is saved
    
    Run the following:
        python extract_top_genes.py -f <PATH_TO_FOLDER/> -r <path_to_result_filename> -t <number_of_top_scorers> -s <path_to_sequences_file>
    
        e.g. 
        python extract_top_genes.py -f Documents/SOY/soy-soy/ -r Documents/SOY/soy_top40 -t 40 -s Documents/SOY/Wm82.a2.v1.protein.aa
        
        Where soy-soy/ contains all gene .csv prediction files
        This will create an Excel with top 40 scorers saved as soy_top40.xlsx under Documents/SOY/
    
Requirements:
    This worked using the following (newer isoforms may also work):
    python 3.7.10
    pandas 0.24.2
    openpyxl 3.0.7
    tqdm 4.59.0

@author: Eric Arezza
Last updated: Nov. 3, 2021
"""

import os
import atexit
import argparse
import traceback
import pandas as pd
import numpy as np
import tqdm
import time
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# DEFINE GENES OF INTEREST TO HIGHLIGHT RED IN EXCEL AND REPORT % FOUND IN TOP SCORERS
GENES_OF_INTEREST = [
    'Glyma.06G093500', 'Glyma.04G091700', 'Glyma.02G213400', 'Glyma.14G181100', 'Glyma.01G023900', 'Glyma.02G040900', 'Glyma.07G220900',
    'Glyma.20G019200', 'Glyma.20G049600', 'Glyma.03G129000', 'Glyma.04G227600', 'Glyma.13G097600', 'Glyma.17G062000', 'Glyma.08G001800',
    'Glyma.15G162300', 'Glyma.09G056100', 'Glyma.05G152000', 'Glyma.08G108800', 'Glyma.18G263000', 'Glyma.11G254700', 'Glyma.18G266800',
    'Glyma.13G073000', 'Glyma.U018700', 'Glyma.05G239400', 'Glyma.03G181900', 'Glyma.08G046500', 'Glyma.05G183500', 'Glyma.10G058000',
    'Glyma.13G144800', 'Glyma.09G205000', 'Glyma.13G129500', 'Glyma.01G018000', 'Glyma.08G019100', 'Glyma.11G080700', 'Glyma.19G182400',
    'Glyma.10G264300', 'Glyma.09G090000', 'Glyma.14G185700', 'Glyma.02G218300', 'Glyma.14G105700', 'Glyma.15G169800', 'Glyma.09G063100',
    'Glyma.13G350500', 'Glyma.05G171400', 'Glyma.05G152000', 'Glyma.08G108800', 'Glyma.11G254700', 'Glyma.12G078400', 'Glyma.08G209500',
    'Glyma.06G263000', 'Glyma.12G078300', 'Glyma.18G054400', 'Glyma.11G160400', 'Glyma.03G031900', 'Glyma.15G024000', 'Glyma.01G136100',
    'Glyma.08G129900', 'Glyma.12G139600', 'Glyma.04G173700', 'Glyma.09G018500', 'Glyma.15G124600', 'Glyma.02G196200', 'Glyma.17G003400',
    'Glyma.07G270600', 'Glyma.12G189000', 'Glyma.13G312700', 'Glyma.13G064800', 'Glyma.12G095100', 'Glyma.10G262600', 'Glyma.12G028800',
    'Glyma.11G103900', 'Glyma.13G314900', 'Glyma.02G195900', 'Glyma.19G199300', 'Glyma.03G202600', 'Glyma.10G180600', 'Glyma.18G040000',
    'Glyma.12G186600', 'Glyma.19G200200', 'Glyma.05G140400', 'Glyma.12G032500', 'Glyma.06G136900', 'Glyma.04G228000', 'Glyma.11G107500',
    'Glyma.06G093500', 'Glyma.04G004000', 'Glyma.06G003600', 'Glyma.05G189100', 'Glyma.05G140400', 'Glyma.03G088500', 'Glyma.08G095800',
    'Glyma.18G040000', 'Glyma.16G084800', 'Glyma.09G228500', 'Glyma.12G008000', 'Glyma.11G216500', 'Glyma.10G258300', 'Glyma.04G091700',
    'Glyma.20G132800', 'Glyma.08G146800', 'Glyma.20G133400', 'Glyma.12G128600', 'Glyma.16G044800', 'Glyma.06G277000', 'Glyma.19G106900',
    'Glyma.02G213400', 'Glyma.06G198400', 'Glyma.14G181100', 'Glyma.03G009500', 'Glyma.01G023900', 'Glyma.17G074700', 'Glyma.08G014900',
    'Glyma.02G040900', 'Glyma.07G071000', 'Glyma.02G203000', 'Glyma.03G026900', 'Glyma.20G049600', 'Glyma.05G208300', 'Glyma.01G140600',
    'Glyma.03G129000', 'Glyma.08G111500', 'Glyma.05G153800', 'Glyma.07G198600', 'Glyma.13G237800', 'Glyma.15G075600', 'Glyma.18G151800',
    'Glyma.13G177800', 'Glyma.08G343800', 'Glyma.02G191200', 'Glyma.03G253000', 'Glyma.19G250600', 'Glyma.16G082800', 'Glyma.03G090700',
    'Glyma.18G074100', 'Glyma.13G314900', 'Glyma.08G332900', 'Glyma.02G302500', 'Glyma.12G186600', 'Glyma.14G011600', 'Glyma.17G220000',
    'Glyma.08G350600', 'Glyma.01G145800', 'Glyma.08G320500', 'Glyma.01G050600', 'Glyma.10G155800', 'Glyma.03G198400', 'Glyma.20G232500',
    'Glyma.18G165200', 'Glyma.19G196300', 'Glyma.07G118700', 'Glyma.10G134000', 'Glyma.16G178800', 'Glyma.20G037900', 'Glyma.15G047500',
    'Glyma.08G185200', 'Glyma.09G131500', 'Glyma.17G182500', 'Glyma.19G098200', 'Glyma.04G128200', 'Glyma.08G032900', 'Glyma.07G152400',
    'Glyma.12G014500', 'Glyma.18G203500', 'Glyma.01G119600', 'Glyma.11G110500', 'Glyma.03G056000', 'Glyma.10G281800', 'Glyma.11G063900',
    'Glyma.20G107500', 'Glyma.02G059000', 'Glyma.16G141700', 'Glyma.01G178300', 'Glyma.15G047500', 'Glyma.01G003800', 'Glyma.19G144800',
    'Glyma.07G128100', 'Glyma.16G097900', 'Glyma.08G185200', 'Glyma.10G155800', 'Glyma.20G232500', 'Glyma.07G072100', 'Glyma.03G011000',
    'Glyma.08G286500', 'Glyma.13G341100', 'Glyma.U027700', 'Glyma.11G134000', 'Glyma.12G058100', 'Glyma.05G207400', 'Glyma.08G014100',
    'Glyma.08G093400', 'Glyma.04G010700', 'Glyma.01G107900', 'Glyma.13G326600', 'Glyma.14G023000', 'Glyma.19G247400', 'Glyma.15G033300',
    'Glyma.02G062700', 'Glyma.13G126600', 'Glyma.20G227500', 'Glyma.19G042300', 'Glyma.17G120900', 'Glyma.05G012900', 'Glyma.06G178800',
    'Glyma.15G169800', 'Glyma.09G063100', 'Glyma.13G350500', 'Glyma.05G040600', 'Glyma.06G076000', 'Glyma.04G075000', 'Glyma.04G200500',
    'Glyma.04G123800', 'Glyma.17G085700', 'Glyma.04G187000', 'Glyma.01G245100', 'Glyma.15G024000', 'Glyma.06G165000', 'Glyma.04G000200',
    'Glyma.06G000100', 'Glyma.19G102000', 'Glyma.11G000300', 'Glyma.12G098900', 'Glyma.06G305700', 'Glyma.12G193800', 'Glyma.13G308700',
    'Glyma.16G049400', 'Glyma.16G147200', 'Glyma.11G080600', 'Glyma.01G162800', 'Glyma.05G035900', 'Glyma.06G178700', 'Glyma.09G035500',
    'Glyma.08G072300', 'Glyma.08G008200', 'Glyma.15G047500', 'Glyma.17G091500', 'Glyma.08G185200', 'Glyma.09G136900', 'Glyma.15G140000',
    'Glyma.08G072200', 'Glyma.16G182300'
    ]

MAX_SHEETS = 200
MAX_COLS = 800

# DEFINE COMMANDLINE ARGUMENTS
describe_help = 'python extract_top_genes.py -f PATH_TO_FOLDER/ -r path_to_result_filename.csv -t 40 -s sequences.fasta'
parser = argparse.ArgumentParser(description=describe_help)
parser.add_argument('-f', '--files', help='Full path to folder with .csv gene files', type=str)
parser.add_argument('-r', '--result', help='Full path to result file', type=str, default=os.getcwd() + '/top_genes.xlsx')
parser.add_argument('-t', '--top', help='Number of top interactors to include', type=int)
parser.add_argument('-s', '--sequences', help='Full path to fasta file containing all sequences', type=str)
parser.add_argument('-a', '--all', help='Flag to all gene isoforms', action='store_true')
parser.add_argument('-pathogen_only', '--pathogen_only', help='Flag to extract only pathogen scores from any soy gene files', action='store_true')
parser.add_argument('-soy_only', '--soy_only', help='Flag to extract only soy scores from any pathogen gene files', action='store_true')
parser.add_argument('-p', '--prefix', help='Prefix of pathogen gene names for searching/filtering', type=str, default='Hetgly')
args = parser.parse_args()

# DEFINE USEFUL FUNCTIONS
def get_isoforms(df_fasta):
    # Format df for easier use columns as ID, SEQUENCE
    df = df_fasta.copy()
    seq_id = df.iloc[::2, :].reset_index(drop=True)
    seq = df.iloc[1::2, :].reset_index(drop=True)
    seq_id.insert(1, 1, seq)
    seq_id[0] = seq_id[0].str.replace('>', '')
    seq_id[0] = seq_id[0].str.replace(r'.p', ' ', regex=True)
    seq_id[0] = [ i[0] for i in seq_id[0].str.split(' ') ]
    df = seq_id.copy()
    print('\t%s total genes in sequences file.'%df.shape[0])
    
    # Sort genes by having single or multi isoforms
    isoforms = np.array([ v[-1] for v in df[0].str.split('.') ])
    df.insert(2, 'Isoform', isoforms)
    df[0] = np.array([ '.'.join(i[:-1]) for i in df[0].str.split('.') ])
    seq_lengths = np.array([ len(i) for i in df[1] ])
    df = pd.DataFrame(data={0: df[0], 1: seq_lengths, 'Isoform': df['Isoform']})
    multi_isoforms = df[df.duplicated(subset=[0], keep=False)]
    multi_isoforms.reset_index(drop=True, inplace=True)
    single_isoforms = df[~df[0].isin(multi_isoforms[0].unique())]
    single_isoforms.reset_index(drop=True, inplace=True)
    
    # Sort multi isoforms by sequence length
    #multi_isoforms = multi_isoforms.sort_values(by=[0,1], ascending=False)
    multi_isoforms = multi_isoforms.sort_values(by=[0,'Isoform'], ascending=True)
    multi_isoforms.reset_index(inplace=True, drop=True)
    
    # Get multi isoforms with same length
    same_length = multi_isoforms[multi_isoforms.duplicated(subset=[0,1], keep=False)]
    same_length.reset_index(drop=True, inplace=True)
    
    # Get longest isoforms from genes with multiple isoforms
    longest = multi_isoforms.drop_duplicates(subset=[0], keep='first')
    
    # Get genes with multiple equally longest isoforms
    many_longest = same_length.merge(longest, on=[0, 1])
    many_longest.drop(columns='Isoform_y', inplace=True)
    many_longest.rename(columns={'Isoform_x': 'Isoform'}, inplace=True)
    many_longest.reset_index(drop=True, inplace=True)
    
    # Combine all longest isoforms and all genes with only one isoform, removing redundant
    isoforms = longest.append(many_longest, ignore_index=True)
    isoforms = isoforms.append(single_isoforms, ignore_index=True)
    isoforms = isoforms.drop_duplicates(subset=[0, 1])
    isoforms.reset_index(drop=True, inplace=True)
    isoforms.sort_values(by=[0, 'Isoform'], inplace=True)
    isoforms.reset_index(drop=True, inplace=True)
    isoforms[0] = isoforms[0] + '.' + isoforms['Isoform']
    
    many_longest[0] = many_longest[0] + '.' + many_longest['Isoform']

    return isoforms[0], many_longest[0]
    
def highlight_genes(gene):
    bg_color = 'pink' if pd.Series([gene]).isin(GENES_OF_INTEREST)[0] == True else 'none'
    return 'background-color: %s' % (bg_color)
def color_genes(gene):
    color = 'red' if pd.Series([gene]).isin(GENES_OF_INTEREST)[0] == True else 'black'
    return 'color: %s' % (color)

def get_saved_genes_sheet(filename, sheetname=None):
    # Start new Excel file if not found
    if sheetname == None:
        sheet = 'Sheet1'
    else:
        sheet = sheetname
    if not os.path.exists(filename):
        return np.array([], dtype=str), sheet
    
    # Get list of gene columns already recorded
    genes_recorded = np.array([], dtype=str)
    excel_file = pd.ExcelFile(filename)
    
    # Start new file if file sheets are maxed out
    if len(excel_file.sheet_names) > MAX_SHEETS:
        args.result = ''.join(args.result[:-1]) + '_new.' + args.result[-1]
    
    # Record saved gene columns
    for sheets in excel_file.sheet_names:
        genes_recorded = np.append(genes_recorded, excel_file.parse(sheets).columns)
    
    # Get current sheetname and number of columns used
    num_cols = excel_file.parse(sheet).shape[1]
    if num_cols < MAX_COLS:
        return genes_recorded, sheet
    else:
        return genes_recorded, "".join(filter(lambda x: not x.isdigit(), sheetname))+str(len(excel_file.sheet_names)+1)
    
def write_to_excel(file, df):
    saved_genes, sheet = get_saved_genes_sheet(file)
    # Create file to append if not existing
    if not os.path.exists(file):
        writer = pd.ExcelWriter(file, engine='openpyxl')
        pd.DataFrame().to_excel(writer, sheet_name=sheet, index=False)
        writer.save()
        writer.close()
        
    # Load file to append
    book = load_workbook(file)
    writer = pd.ExcelWriter(file, engine='openpyxl', mode='a')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    # Add SOY gene columns first
    #pbar = tqdm.tqdm(total=np.array(df[[ x for x in df.columns if 'Glyma' in x ]]).shape[0])
    pbar = tqdm.tqdm(total=df.columns.isin(saved_genes).shape[0] - df.columns.isin(saved_genes).sum())
    for i in range(0, 21):
        to_write, sheet = get_df_chromosome(df, num=i)
        if to_write.empty:
            continue
        excel = to_write.style.applymap(highlight_genes)
        excel = excel.applymap(color_genes)
        excel.to_excel(writer, sheet_name=sheet, index=False)
        saved_genes = np.append(saved_genes, to_write.columns)
        Worksheet(book, title=sheet)
        pbar.update(to_write.shape[1])
        df.drop(columns=to_write.columns, inplace=True)
        
    writer.save()
    writer.close()
    #pbar.close()
    
    # Load file to append
    book = load_workbook(file)
    writer = pd.ExcelWriter(file, engine='openpyxl', mode='a')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    # Add pathogen gene columns until completed
    cols = 0
    sheet_num = 1
    sheet = 'Sheet1'
    #pbar = tqdm.tqdm(total=df.columns.isin(saved_genes).shape[0] - df.columns.isin(saved_genes).sum())
    while not all(df.columns.isin(saved_genes)):
        to_write = df[df.columns[cols:cols+MAX_COLS]]
        #to_write = get_df_chromosome(df, num=sheet_num)
        excel = to_write.style.applymap(highlight_genes)
        excel = excel.applymap(color_genes)
        excel.to_excel(writer, sheet_name=sheet, index=False)
        saved_genes = np.append(saved_genes, to_write.columns)
        sheet = "".join(filter(lambda x: not x.isdigit(), sheet))+str(sheet_num+1)
        Worksheet(book, title=sheet)
        pbar.update(to_write.shape[1])
        sheet_num += 1
        cols += MAX_COLS
        
    writer.save()
    writer.close()
    pbar.close()
    
def get_df_chromosome(df, num=0):
    if num == 0:
        chromosome_soy = df[[ c for c in df.columns if 'Glyma.U' in c ]]
        sheet = 'Soy Chromosome U'
    else:
        chromosome_soy = df[[ c for c in df.columns if ('Glyma.' +('%s'%num).zfill(2) + 'G') in c ]]
        sheet = 'Soy Chromosome %s'%num
    
    return chromosome_soy, sheet

def show_duration(t_start):
    # Display duration of run
    t_duration = time.time() - t_start
    day = t_duration // (24 * 3600)
    t_duration = t_duration % (24 * 3600)
    hour = t_duration // 3600
    t_duration %= 3600
    minutes = t_duration // 60
    t_duration %= 60
    seconds = t_duration
    print('Days', day)
    print('Hours', hour)
    print('Minutes', minutes)
    print('Seconds', seconds)

# MAIN RUN
if __name__ == '__main__':
    t_start = time.time()
    # Show args for double-checking input
    print(args)
    
    try:
    
        # If only want to get non-Soy gene scores from files
        if args.pathogen_only:
            files = np.array([ f for f in os.listdir(path=args.files) if '.csv' in f and 'Glyma' in f ])
        # If only want to get soy scores from files
        elif args.soy_only:
            files = np.array([ f for f in os.listdir(path=args.files) if '.csv' in f and args.prefix in f ])
        else:
            # Get all gene files in folder
            files = np.array([ f for f in os.listdir(path=args.files) if '.csv' in f and ('Glyma' in f or args.prefix in f) ])
        print('\nNumber of gene files:', len(files))
        
        # Skip any saved gene columns if exist already
        saved_genes, sheetname = get_saved_genes_sheet(args.result)
        print('%s genes already saved'%saved_genes.shape[0])
        
        if not args.all:
            print('Getting longest sequenced isoforms...')
            seq = pd.read_csv(args.sequences, sep='\n', header=None)
            isoforms, many_longest = get_isoforms(seq)
            print('\t%s relevant gene isoforms\n\t%s have multiple equally long sequences...'%(isoforms.shape[0], isoforms[isoforms.isin(many_longest)].shape[0]))
            # Only process gene files for relevant isoforms
            #files = np.array([ i for i in files if i.replace('.csv', '' ) in isoforms.values or i.replace('.csv', '' ) in many_longest.values ])
            files = np.array([ i for i in files if i.replace('.csv', '' ) in isoforms.values and i.replace('.csv', '' ) not in saved_genes ])
            print('\n%s relevant files to process...\n'%len(files))
        
        # For result file
        final = pd.DataFrame()
        final_isoforms = pd.DataFrame()
        
        # Iterate through each file
        print('\nIterating through files...')
        for f in tqdm.tqdm(files):
    
            # Get gene name from filename
            gene = f.replace('.csv', '')
            
            # Read file
            df = pd.read_csv(args.files + f)
            
            # Remove any genes that are not SOY or pathogen
            df = df[(df[df.columns[1]].str.contains('Glyma')) | (df[df.columns[1]].str.contains(args.prefix))]
            df.reset_index(drop=True, inplace=True)
            
            # If only want to consider non-Soy genes for interactor scores
            if args.pathogen_only:
                df = df[df[df.columns[1]].str.contains(args.prefix)]
                df.reset_index(drop=True, inplace=True)
            if args.soy_only:
                df = df[df[df.columns[1]].str.contains('Glyma')]
                df.reset_index(drop=True, inplace=True)
            
            # Sort by scores in descending order
            df.sort_values(by=df.columns[-1], ascending=False, inplace=True)
            df.reset_index(drop=True, inplace=True)
            
            # Keep copy for recording interactor isoform number
            df_isoforms = df.copy()
            # Remove gene isoform number (after the last '.')
            df[df.columns[1]] = [ '.'.join(g[:-1]) for g in df[df.columns[1]].str.split('.').values ]
            
            # Drop any duplicated interactor genes from list (removes isoform variants, keeping the top scored one)
            df = df.drop_duplicates(subset=[df.columns[1]])
            
            # Re-sort non-duplicated genes
            df.sort_values(by=df.columns[-1], ascending=False, inplace=True)
            
            # Get top interactors
            df = df.iloc[:args.top]
            df_isoforms = df_isoforms.iloc[df.index]
            
            df.reset_index(drop=True, inplace=True)
            df_isoforms.reset_index(drop=True, inplace=True)

            # Count % of top interactors found in genes of interest
            percent_interested = ( df[df.columns[1]].isin(GENES_OF_INTEREST).sum() / args.top )*100
            
            # Create column for gene and include % of top in genes of interest
            gene_info = pd.DataFrame(data=df[df.columns[1]].values, columns=[gene])
            gene_info_isoforms = pd.DataFrame(data=df_isoforms[df_isoforms.columns[1]].values, columns=[gene])
            
            gene_info = gene_info[gene].append(pd.Series(percent_interested), ignore_index=True)
            gene_info_isoforms = gene_info_isoforms[gene].append(pd.Series(percent_interested), ignore_index=True)
            
            gene_info = pd.DataFrame(gene_info, columns=[gene])
            gene_info_isoforms = pd.DataFrame(gene_info_isoforms, columns=[gene])
    
            #Add gene info to final
            final.insert(len(final.columns), gene, gene_info[gene])
            final_isoforms.insert(len(final_isoforms.columns), gene, gene_info_isoforms[gene])
        
        final = final[sorted(final.columns)]
        final_isoforms = final_isoforms[sorted(final_isoforms.columns)]
        show_duration(t_start)
        
        if final.empty:
            print('No columns to add...\nDone!\n')
            show_duration(t_start)
            exit()
        
        # Write to excel
        print('Creating coloured Excel, %s gene columns...'%final.shape[1])
        write_to_excel(args.result, final)
        
        show_duration(t_start)
        
        print('Creating Excel with isoform numbers, %s gene columns...'%final_isoforms.shape[1])
        iso_filename = args.result.split('.')
        iso_filename = ''.join(iso_filename[:-1]) + '_isoform_numbers.' + iso_filename[-1]
        write_to_excel(iso_filename, final_isoforms)
        
        print('Done!\n')
        show_duration(t_start)
    except (KeyboardInterrupt, Exception):
        atexit.register(show_duration, t_start)
        traceback.print_exc()
    
